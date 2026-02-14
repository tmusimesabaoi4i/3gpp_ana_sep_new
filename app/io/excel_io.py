"""
io/excel_io.py  –  SELECT → Excel 出力 / 多シート統合

機能:
  1. export_select  : 単一 SELECT → Excel
  2. merge_csvs_to_excel : 複数 CSV → 1 Excel (シート名=キー)
  3. build_analysis_excel : A〜E CSV → ALL_*/CO_* シート + META シート
"""
from __future__ import annotations

import csv as csv_mod
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from app.core.types import NullPolicySpec, SelectSpec


class ExcelIO:
    """Excel 出力マネージャ"""

    @staticmethod
    def export_select(
        sio: "SqliteIO",
        select: SelectSpec,
        out_path: str | Path,
        null_policy: NullPolicySpec | None = None,
        chunk_size: int = 5000,
    ) -> int:
        try:
            from openpyxl import Workbook
        except ImportError:
            raise RuntimeError(
                "Excel 出力には openpyxl が必要です。pip install openpyxl を実行してください。"
            )

        from app.io.sqlite_io import SqliteIO

        out = Path(out_path)
        out.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title="data")
        ws.append(select.columns)

        total = 0
        for chunk in sio.query_iter(select.sql, select.params, chunk_size):
            for row in chunk:
                ws.append(_apply_null_policy(row, null_policy))
                total += 1

        wb.save(str(out))
        return total


def _apply_null_policy(row: tuple, policy: NullPolicySpec | None) -> list[Any]:
    if policy is None:
        return list(row)
    return [
        (policy.text_null if policy.text_null is not None else "")
        if v is None else v
        for v in row
    ]


def _try_numeric(val: str) -> Any:
    """CSV文字列を数値に変換（可能なら）"""
    if not val:
        return val
    try:
        if "." in val:
            return float(val)
        return int(val)
    except (ValueError, TypeError):
        return val


def _safe_sheet_name(name: str) -> str:
    """Excel のシート名制約 (最大31文字, 無効文字除去)"""
    name = re.sub(r'[\\/*?\[\]:]', '_', name)
    return name[:31]


def merge_csvs_to_excel(
    csv_map: dict[str, str | Path],
    excel_path: str | Path,
) -> None:
    """複数CSVを1つのExcelファイル（複数シート）に統合する。"""
    try:
        from openpyxl import Workbook
    except ImportError:
        raise RuntimeError("Excel出力には openpyxl が必要です。pip install openpyxl")

    out = Path(excel_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    if wb.sheetnames:
        del wb[wb.sheetnames[0]]

    for sheet_name, csv_path in csv_map.items():
        csv_path = Path(csv_path)
        if not csv_path.exists():
            continue
        ws = wb.create_sheet(title=_safe_sheet_name(sheet_name))
        with open(csv_path, "r", encoding="utf-8-sig") as f:
            reader = csv_mod.reader(f)
            for i, row in enumerate(reader):
                if i == 0:
                    ws.append(row)
                else:
                    ws.append([_try_numeric(v) for v in row])

    wb.save(str(out))


def build_analysis_excel(
    csv_map: dict[str, str | Path],
    excel_path: str | Path,
    companies: dict[str, str],
    job_meta: list[dict[str, Any]],
    include_meta: bool = True,
) -> Path:
    """A〜E 分析 CSV → ALL_*/CO_*_* シート分割 + META シートの Excel を生成する。

    Args:
        csv_map: {analysis_key: csv_path}  例: {"A_filing_ts": "out/A_filing_ts.csv"}
        excel_path: 出力 Excel パス
        companies: {display_key: LIKE_pattern} 例: {"NTT_DOCOMO": "DOCOMO"}
        job_meta: 各ジョブのメタ情報 [{job_id, template, scope_summary, unique_unit, period}]
        include_meta: META シートを含めるか

    Returns:
        生成した Excel のパス
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError("Excel出力には openpyxl が必要です。pip install openpyxl")

    out = Path(excel_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    if wb.sheetnames:
        del wb[wb.sheetnames[0]]

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    sheets_created: list[dict[str, str]] = []  # for META

    for analysis_key, csv_path in csv_map.items():
        csv_path = Path(csv_path)
        if not csv_path.exists():
            continue

        # CSV 全行読み込み
        rows: list[list[str]] = []
        with open(csv_path, "r", encoding="utf-8-sig") as f:
            reader = csv_mod.reader(f)
            for row in reader:
                rows.append(row)

        if len(rows) < 2:
            continue

        header = rows[0]
        data_rows = rows[1:]

        # company 列のインデックスを検出
        company_col_idx = _find_col_index(header, "company")

        # ── ALL シート ──
        all_sheet_name = _safe_sheet_name(f"ALL_{analysis_key}")
        ws_all = wb.create_sheet(title=all_sheet_name)
        _write_header(ws_all, header, header_font, header_fill)
        for row in data_rows:
            ws_all.append([_try_numeric(v) for v in row])
        sheets_created.append({
            "sheet": all_sheet_name,
            "analysis": analysis_key,
            "filter": "ALL (全データ)",
        })

        # ── CO_<key> シート（企業ごとにフィルタ） ──
        if company_col_idx is not None and companies:
            for display_key, like_pattern in companies.items():
                co_sheet_name = _safe_sheet_name(f"CO_{display_key}_{analysis_key}")
                filtered = [
                    row for row in data_rows
                    if _company_match(row, company_col_idx, like_pattern)
                ]
                if not filtered:
                    continue
                ws_co = wb.create_sheet(title=co_sheet_name)
                _write_header(ws_co, header, header_font, header_fill)
                for row in filtered:
                    ws_co.append([_try_numeric(v) for v in row])
                sheets_created.append({
                    "sheet": co_sheet_name,
                    "analysis": analysis_key,
                    "filter": f"company LIKE '%{like_pattern}%'",
                })

    # ── META シート ──
    if include_meta:
        ws_meta = wb.create_sheet(title="META", index=0)
        _write_meta_sheet(ws_meta, job_meta, sheets_created, header_font, header_fill)

    wb.save(str(out))
    return out


def _find_col_index(header: list[str], col_name: str) -> Optional[int]:
    """ヘッダから列名を探す（大文字小文字無視）"""
    lower = col_name.lower()
    for i, h in enumerate(header):
        if h.strip().lower() == lower:
            return i
    return None


def _company_match(row: list[str], col_idx: int, pattern: str) -> bool:
    """行の company 列が LIKE pattern に部分一致するか"""
    if col_idx >= len(row):
        return False
    return pattern.upper() in row[col_idx].upper()


def _write_header(ws: Any, header: list[str], font: Any, fill: Any) -> None:
    """ヘッダ行を書式付きで書き込む"""
    ws.append(header)
    for cell in ws[1]:
        cell.font = font
        cell.fill = fill


def _write_meta_sheet(
    ws: Any,
    job_meta: list[dict[str, Any]],
    sheets_created: list[dict[str, str]],
    font: Any,
    fill: Any,
) -> None:
    """META シートを書き込む"""
    # ── セクション1: ジョブ一覧 ──
    ws.append(["ISLD Analysis Results — META"])
    ws.append([f"生成日時: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws.append([])
    ws.append(["=== Job 一覧 ==="])
    meta_header = ["job_id", "template", "job_description", "scope_summary",
                    "unique_unit", "period"]
    ws.append(meta_header)
    for cell in ws[ws.max_row]:
        cell.font = font
        cell.fill = fill
    for m in job_meta:
        ws.append([
            m.get("job_id", ""),
            m.get("template", ""),
            m.get("job_description", ""),
            m.get("scope_summary", ""),
            m.get("unique_unit", ""),
            m.get("period", ""),
        ])

    ws.append([])
    ws.append(["=== Sheet 一覧 ==="])
    sheet_header = ["シート名", "分析キー", "フィルタ条件"]
    ws.append(sheet_header)
    for cell in ws[ws.max_row]:
        cell.font = font
        cell.fill = fill
    for s in sheets_created:
        ws.append([s["sheet"], s["analysis"], s["filter"]])
