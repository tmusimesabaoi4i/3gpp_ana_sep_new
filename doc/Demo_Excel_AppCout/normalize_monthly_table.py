#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
normalize_monthly_table.py

【用途】
会社ごとに 3列セット（出願人名 / 月次日付 / 件数）が横に並んでいるExcelを読み、
「1990-01-01 から 1か月刻み」でラベルを統一した
"年月 × 企業" の集計テーブル（欠損月は0）を Excel に出力します。

入力イメージ（1行目に企業名、以降は [name, date, count] が繰り返し）:
  Ericsson | (date) | (count) | Huawei | (date) | (count) | NEC | (date) | (count) | ...

出力イメージ:
  年月, Ericsson, Huawei, NEC, DOCOMO, Toyota
  1990-01-01, 2, 0, 0, 0, 0
  1990-02-01, 1, 0, 2, 0, 0
  ...

【実行例】
  python normalize_monthly_table.py --input in.xlsx --output out.xlsx
  python normalize_monthly_table.py --input in.xlsx --output out.xlsx --start 1990-01-01
  python normalize_monthly_table.py --input in.xlsx --output out.xlsx --sheet Sheet1
  python normalize_monthly_table.py --input in.xlsx --output out.xlsx --csv-out out.csv
"""

from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Union

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


DateLike = Union[date, datetime]


@dataclass(frozen=True)
class Block:
    company: str
    start_col: int  # name_col=start_col, date_col=start_col+1, count_col=start_col+2


def _to_month_start(d: date) -> date:
    return date(d.year, d.month, 1)


def _parse_month_cell(v: object) -> Optional[date]:
    """
    Excelセルの値を date(YYYY-MM-01) に変換する。
    受け入れ: datetime/date, "YYYY-MM-DD", "YYYY/MM/DD", "YYYY-MM", "YYYY/MM"
    """
    if v is None:
        return None
    if isinstance(v, datetime):
        return _to_month_start(v.date())
    if isinstance(v, date):
        return _to_month_start(v)

    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None

        fmts = ["%Y-%m-%d", "%Y/%m/%d", "%Y-%m", "%Y/%m"]
        for fmt in fmts:
            try:
                dt = datetime.strptime(s, fmt)
                return date(dt.year, dt.month, 1)
            except ValueError:
                continue
        return None

    # 数値シリアル日付などは openpyxl が通常 datetime にしてくれるが、
    # 念のためここでは扱わない
    return None


def _parse_int_cell(v: object) -> Optional[int]:
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        if not s:
            return None
        try:
            return int(float(s))
        except ValueError:
            return None
    return None


def _month_range(start: date, end: date) -> List[date]:
    """
    start〜end（両端含む）を月初dateで生成
    """
    start = _to_month_start(start)
    end = _to_month_start(end)
    out: List[date] = []
    y, m = start.year, start.month
    while (y, m) <= (end.year, end.month):
        out.append(date(y, m, 1))
        m += 1
        if m == 13:
            y += 1
            m = 1
    return out


def _detect_blocks(ws) -> List[Block]:
    """
    1行目を見て、値が入っているセルを company ブロック開始列として扱う。
    ブロックは 3列固定（name/date/count）とする。
    """
    max_col = ws.max_column
    blocks: List[Block] = []
    for c in range(1, max_col + 1):
        v = ws.cell(1, c).value
        if isinstance(v, str):
            name = v.strip()
            if name:
                # 3列確保できるものだけ採用
                if c + 2 <= max_col:
                    blocks.append(Block(company=name, start_col=c))
    return blocks


def normalize_excel(
    input_xlsx: Path,
    output_xlsx: Path,
    *,
    sheet_name: Optional[str] = None,
    start_date: date = date(1990, 1, 1),
    csv_out: Optional[Path] = None,
    rename_map: Optional[Dict[str, str]] = None,
) -> None:
    wb = load_workbook(input_xlsx, data_only=True, read_only=True)
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    blocks = _detect_blocks(ws)
    if not blocks:
        raise RuntimeError("企業ブロックが検出できませんでした（1行目に企業名がある想定です）。")

    # 企業名リネーム（例: "Ericson" -> "Ericsson"）を適用した表示名リスト
    def disp_company(x: str) -> str:
        if rename_map and x in rename_map:
            return rename_map[x]
        return x

    companies = [disp_company(b.company) for b in blocks]

    # 集計: (month, company) -> count_sum
    agg: Dict[Tuple[date, str], int] = {}
    max_month: Optional[date] = None

    # データ読み（2行目以降）
    for r in range(2, ws.max_row + 1):
        for b in blocks:
            name_col = b.start_col
            date_col = b.start_col + 1
            count_col = b.start_col + 2

            # nameは基本不要だが、完全空行判定に使える
            name_v = ws.cell(r, name_col).value
            date_v = ws.cell(r, date_col).value
            count_v = ws.cell(r, count_col).value

            if name_v is None and date_v is None and count_v is None:
                continue

            month = _parse_month_cell(date_v)
            cnt = _parse_int_cell(count_v)
            if month is None or cnt is None:
                continue

            comp = disp_company(b.company)
            key = (month, comp)
            agg[key] = agg.get(key, 0) + cnt

            if max_month is None or month > max_month:
                max_month = month

    if max_month is None:
        raise RuntimeError("有効な (date, count) データが見つかりませんでした。")

    # 1990年開始で統一（指定があればそれ）
    end_date = max_month
    if end_date < start_date:
        end_date = start_date

    months = _month_range(start_date, end_date)

    # 出力ワークブック作成
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "monthly_table"

    header = ["年月"] + companies
    out_ws.append(header)
    for cell in out_ws[1]:
        cell.font = Font(bold=True)

    for m in months:
        row = [m.isoformat()]
        for comp in companies:
            row.append(agg.get((m, comp), 0))
        out_ws.append(row)

    # 見やすさ調整
    out_ws.freeze_panes = "A2"
    out_ws.column_dimensions["A"].width = 12
    for i in range(2, len(header) + 1):
        out_ws.column_dimensions[get_column_letter(i)].width = max(10, len(header[i - 1]) + 2)

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(output_xlsx)

    # CSVも欲しい場合
    if csv_out is not None:
        csv_out.parent.mkdir(parents=True, exist_ok=True)
        with csv_out.open("w", encoding="utf-8", newline="") as f:
            w = csv.writer(f, lineterminator="\n")
            w.writerow(header)
            for m in months:
                row = [m.isoformat()] + [agg.get((m, comp), 0) for comp in companies]
                w.writerow(row)


def main() -> int:
    ap = argparse.ArgumentParser(description="企業ごとの月次データを 1990年〜月次で揃えたテーブルに正規化します。")
    ap.add_argument("--input", required=True, help="入力Excel (.xlsx)")
    ap.add_argument("--output", required=True, help="出力Excel (.xlsx)")
    ap.add_argument("--sheet", default=None, help="入力シート名（省略時は先頭シート）")
    ap.add_argument("--start", default="1990-01-01", help="開始年月日（既定: 1990-01-01）")
    ap.add_argument("--csv-out", default=None, help="CSVも出力したい場合のパス（任意）")
    ap.add_argument(
        "--rename-ericson",
        action="store_true",
        help='ヘッダの "Ericson" を "Ericsson" に修正して出力する（任意）',
    )

    args = ap.parse_args()

    start_dt = _parse_month_cell(args.start)
    if start_dt is None:
        raise SystemExit(f"[ERROR] --start の形式が不正です: {args.start} (例: 1990-01-01)")

    rename_map = {"Ericson": "Ericsson"} if args.rename_ericson else None

    normalize_excel(
        input_xlsx=Path(args.input),
        output_xlsx=Path(args.output),
        sheet_name=args.sheet,
        start_date=start_dt,
        csv_out=Path(args.csv_out) if args.csv_out else None,
        rename_map=rename_map,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
