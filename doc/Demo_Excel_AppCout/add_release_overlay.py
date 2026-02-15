#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
add_release_overlay.py

【目的】
「年月 × 企業」の月次テーブルに、3GPP Release の期間（Start〜End/Closure）を
“月次ラベル(YYYY-MM-01)” に丸めて重ねて分析できる列を追加したExcelを作る。

【Release表の扱い（重要）】
- Start date は必須（ある前提）
- End date がある場合はそれを使用
- End date が空で Closure date がある場合は Closure date を終端として使用
- End/Closure どちらも空の場合は「終端なし」として、入力テーブル最終月まで有効とみなす

日付文字列は "2027-06-18 (SA#116)" のように括弧があってもOK（先頭のYYYY-MM-DDを抜き出す）

【実行例】
  python add_release_overlay.py --input monthly.xlsx --output monthly_with_release.xlsx
  python add_release_overlay.py --input monthly.xlsx --output monthly_with_release.xlsx --sheet monthly_table
"""

from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


DateLike = Union[date, datetime]


def _to_month_start(d: date) -> date:
    return date(d.year, d.month, 1)


def _parse_month_from_any(v: Any) -> Optional[date]:
    """
    Excelセル値/文字列から month start (YYYY-MM-01) を得る。
    受け入れ:
      - date/datetime
      - "YYYY-MM-DD", "YYYY/MM/DD"
      - "YYYY-MM", "YYYY/MM"
      - "YYYY-MM-DD (SA#xxx)" のような注釈付き（先頭の日付を抽出）
    """
    if v is None:
        return None
    if isinstance(v, datetime):
        return _to_month_start(v.date())
    if isinstance(v, date):
        return _to_month_start(v)
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None

        # 注釈付き "2027-06-18 (SA#116)" など → まず日付部分抽出
        m = re.search(r"(\d{4})-(\d{2})-(\d{2})", s)
        if m:
            y, mo, _d = int(m.group(1)), int(m.group(2)), int(m.group(3))
            return date(y, mo, 1)

        # 月だけ "2027-03" など
        m2 = re.search(r"(\d{4})-(\d{2})", s)
        if m2:
            y, mo = int(m2.group(1)), int(m2.group(2))
            return date(y, mo, 1)

        # スラッシュ
        m3 = re.search(r"(\d{4})/(\d{2})/(\d{2})", s)
        if m3:
            y, mo, _d = int(m3.group(1)), int(m3.group(2)), int(m3.group(3))
            return date(y, mo, 1)

        m4 = re.search(r"(\d{4})/(\d{2})", s)
        if m4:
            y, mo = int(m4.group(1)), int(m4.group(2))
            return date(y, mo, 1)

        return None
    return None


def _as_int(v: Any) -> int:
    if v is None:
        return 0
    if isinstance(v, bool):
        return 0
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        if not s:
            return 0
        try:
            return int(float(s))
        except ValueError:
            return 0
    return 0


@dataclass(frozen=True)
class ReleaseRow:
    code: str          # Rel-19, R99, UMTS, Ph1...
    name: str          # Release 19, Release 1999, UMTS...
    status: str        # Open/Frozen/Closed
    start_raw: str
    end_raw: str
    closure_raw: str
    start_month: date
    end_month_effective: Optional[date]   # None = open ended
    end_source: str                       # "end" / "closure" / "open"


def _build_release_master() -> List[ReleaseRow]:
    """
    ユーザ提示の表をそのまま埋め込む。
    """
    rows = [
        # code, name, status, start, end, closure
        ("Rel-21", "Release 21", "Open",   "2025-11-04", "",                    ""),
        ("Rel-20", "Release 20", "Open",   "2024-03-14", "2027-06-18 (SA#116)", ""),
        ("Rel-19", "Release 19", "Frozen", "2021-06-18", "2025-12-12 (SA#110)", ""),
        ("Rel-18", "Release 18", "Frozen", "2019-09-16", "2024-06-21 (SA#104)", ""),
        ("Rel-17", "Release 17", "Frozen", "2018-06-15", "2022-06-10 (SA#96)",  ""),
        ("Rel-16", "Release 16", "Frozen", "2017-03-22", "2020-07-03 (SA#88-e)",""),
        ("Rel-15", "Release 15", "Frozen", "2016-06-01", "2019-06-07 (SA#84)",  ""),
        ("Rel-14", "Release 14", "Frozen", "2014-09-17", "2017-06-09 (SA#76)",  ""),
        ("Rel-13", "Release 13", "Frozen", "2012-09-30", "2016-03-11 (SA#71)",  ""),
        ("Rel-12", "Release 12", "Frozen", "2011-06-26", "2015-03-13 (SA#67)",  ""),
        ("Rel-11", "Release 11", "Frozen", "2010-01-22", "2013-03-06 (SA#59)",  ""),
        ("Rel-10", "Release 10", "Frozen", "2009-01-20", "2011-06-08 (SA#52)",  ""),
        ("Rel-9",  "Release 9",  "Frozen", "2008-03-06", "2010-03-25 (SA#47)",  ""),
        ("Rel-8",  "Release 8",  "Frozen", "2006-01-23", "2009-03-12 (SA#43)",  ""),
        ("Rel-7",  "Release 7",  "Closed", "2003-10-06", "2008-03-13 (SA#39)",  "2014-09-17 (SA#65)"),
        ("Rel-6",  "Release 6",  "Closed", "2000-03-28", "2005-09-28 (SA#29)",  "2014-09-17 (SA#65)"),
        ("Rel-5",  "Release 5",  "Closed", "2000-05-01", "2002-09-12 (SA-#17)", "2014-09-17 (SA#65)"),
        ("Rel-4",  "Release 4",  "Closed", "1998-08-01", "2001-06-21 (SA-#12)", "2014-09-17 (SA#65)"),
        ("R00",    "Release 2000","Closed","1999-03-30", "",                    "1999-12-17 (SA-#6)"),
        ("R99",    "Release 1999","Closed","1996-11-01", "1999-12-17 (SA-#6)",  "2008-06-05 (SA#40)"),
        ("UMTS",   "UMTS",       "Closed","1994-08-18", "1999-02-12 (SMG-#28)", "2005-06-08 (SA#28)"),
        ("R98",    "Release 1998","Closed","1996-03-26", "1999-02-12 (SMG-#28)", "2005-06-08 (SA#28)"),
        ("R97",    "Release 1997","Closed","1996-04-15", "",                    "2005-06-08 (SA#28)"),
        ("R96",    "Phase 2+",   "Closed","1995-07-20", "",                    "2004-09-16 (SA#25)"),
        ("Ph2",    "Phase 2",    "Closed","1993-03-25", "",                    "2004-09-16 (SA#25)"),
        ("Ph1-EXT","Phase 1 extension","Closed","1994-10-01","",               "2004-09-16 (SA#25)"),
        ("Ph1-DCS","Phase 1 DCS-1800","Closed","1992-12-03","",                "2004-09-16 (SA#25)"),
        ("Ph1",    "Phase 1",    "Closed","1982-12-07", "",                    "2004-09-16 (SA#25)"),
    ]

    out: List[ReleaseRow] = []
    for code, name, status, start_raw, end_raw, closure_raw in rows:
        sm = _parse_month_from_any(start_raw)
        if sm is None:
            raise RuntimeError(f"Start date parse failed: {code} start={start_raw}")

        em = _parse_month_from_any(end_raw) if end_raw else None
        cm = _parse_month_from_any(closure_raw) if closure_raw else None

        if em is not None:
            end_eff = em
            end_source = "end"
        elif cm is not None:
            end_eff = cm
            end_source = "closure"
        else:
            end_eff = None
            end_source = "open"

        out.append(
            ReleaseRow(
                code=code,
                name=name,
                status=status,
                start_raw=start_raw,
                end_raw=end_raw,
                closure_raw=closure_raw,
                start_month=sm,
                end_month_effective=end_eff,
                end_source=end_source,
            )
        )
    return out


def _read_monthly_table(ws) -> Tuple[List[str], List[Tuple[date, List[int]]]]:
    """
    入力: 1行目ヘッダ ["年月", 企業1, 企業2, ...]
         2行目以降 ["YYYY-MM-01", 数値...]
    """
    if ws.max_row < 2 or ws.max_column < 2:
        raise RuntimeError("入力シートが小さすぎます（年月+企業列の形を想定）。")

    header: List[str] = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        header.append(str(v).strip() if v is not None else "")

    data: List[Tuple[date, List[int]]] = []
    for r in range(2, ws.max_row + 1):
        m = _parse_month_from_any(ws.cell(r, 1).value)
        if m is None:
            continue
        vals = [_as_int(ws.cell(r, c).value) for c in range(2, ws.max_column + 1)]
        data.append((m, vals))

    if not data:
        raise RuntimeError("有効な年月行が見つかりませんでした。")

    data.sort(key=lambda x: x[0])
    return header, data


def _write_table(ws, header: List[str], rows: List[List[Any]], freeze: str = "A2") -> None:
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)

    ws.freeze_panes = freeze

    # 幅調整（雑に）
    for i, h in enumerate(header, start=1):
        w = max(10, min(40, len(str(h)) + 2))
        ws.column_dimensions[get_column_letter(i)].width = w


def add_release_overlay(input_xlsx: Path, output_xlsx: Path, sheet_name: Optional[str] = None) -> None:
    releases = _build_release_master()

    wb_in = load_workbook(input_xlsx, data_only=True, read_only=True)
    ws_in = wb_in[sheet_name] if sheet_name else wb_in.active

    header, data = _read_monthly_table(ws_in)
    companies = header[1:]

    min_month = data[0][0]
    max_month = data[-1][0]

    # Releaseの終端なしは、月次テーブルの最終月まで有効とする
    end_by_code: Dict[str, date] = {}
    for r in releases:
        end_by_code[r.code] = r.end_month_effective if r.end_month_effective is not None else max_month

    # イベント（Start/End）を月で引けるように
    start_events: Dict[date, List[ReleaseRow]] = {}
    end_events: Dict[date, List[ReleaseRow]] = {}
    for r in releases:
        start_events.setdefault(r.start_month, []).append(r)
        end_events.setdefault(end_by_code[r.code], []).append(r)

    # 追加列
    extra_cols = [
        "Active_Releases",       # 月に有効なReleaseコード一覧（カンマ）
        "Active_Release_Names",  # 月に有効なRelease名一覧（カンマ）
        "Start_Releases",        # その月が開始月のRelease
        "End_Releases",          # その月が終了月(End/Closure/Max)のRelease
    ]

    # Release別フラグ列（0/1）
    # 例: "Rel-19_ACTIVE", "Rel-19_START", "Rel-19_END"
    per_rel_cols: List[str] = []
    for r in releases:
        per_rel_cols.extend([f"{r.code}_ACTIVE", f"{r.code}_START", f"{r.code}_END"])

    header2 = ["年月"] + companies + extra_cols + per_rel_cols

    # 出力WB
    wb_out = Workbook()

    # sheet1: monthly_table
    ws1 = wb_out.active
    ws1.title = "monthly_table"
    rows1 = [[m.isoformat()] + vals for m, vals in data]
    _write_table(ws1, header, rows1)

    # sheet2: monthly_with_release
    ws2 = wb_out.create_sheet("monthly_with_release")
    rows2: List[List[Any]] = []

    for m, vals in data:
        active_codes: List[str] = []
        active_names: List[str] = []
        per_flags: List[int] = []

        # start/end その月に該当するリスト（文字列用）
        starts = start_events.get(m, [])
        ends = end_events.get(m, [])

        for r in releases:
            endm = end_by_code[r.code]
            is_active = int(r.start_month <= m <= endm)
            is_start = int(m == r.start_month)
            is_end = int(m == endm)
            if is_active:
                active_codes.append(r.code)
                active_names.append(f"{r.name}({r.status})")
            per_flags.extend([is_active, is_start, is_end])

        row = (
            [m.isoformat()]
            + vals
            + [
                ", ".join(active_codes),
                ", ".join(active_names),
                ", ".join([x.code for x in starts]),
                ", ".join([x.code for x in ends]),
            ]
            + per_flags
        )
        rows2.append(row)

    _write_table(ws2, header2, rows2)

    # sheet3: release_master
    ws3 = wb_out.create_sheet("release_master")
    header3 = [
        "code", "name", "status",
        "start_raw", "start_month",
        "end_raw", "closure_raw",
        "end_month_effective", "end_source"
    ]
    rows3: List[List[Any]] = []
    for r in releases:
        end_eff = end_by_code[r.code]
        rows3.append([
            r.code, r.name, r.status,
            r.start_raw, r.start_month.isoformat(),
            r.end_raw, r.closure_raw,
            end_eff.isoformat(),
            r.end_source,
        ])
    _write_table(ws3, header3, rows3, freeze="A2")

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(output_xlsx)


def main() -> int:
    ap = argparse.ArgumentParser(description="月次テーブルに3GPP Release期間(Start〜End/Closure)の重ね合わせ列を追加します。")
    ap.add_argument("--input", required=True, help="入力Excel (.xlsx)")
    ap.add_argument("--output", required=True, help="出力Excel (.xlsx)")
    ap.add_argument("--sheet", default=None, help="入力シート名（省略時は先頭シート）")
    args = ap.parse_args()

    add_release_overlay(Path(args.input), Path(args.output), sheet_name=args.sheet)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
